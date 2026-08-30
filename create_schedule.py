import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import random

# Create workbook
wb = Workbook()

# Remove default sheet
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# Create main schedule sheet
ws_schedule = wb.create_sheet('일정관리', 0)

# Headers
headers = ['대분류', '중분류', '프로그램ID', '작업명', '작업자', '시작일자', '종료일자', '실제근무일']
for col_num, header in enumerate(headers, 1):
    cell = ws_schedule.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Set column widths
ws_schedule.column_dimensions['A'].width = 15
ws_schedule.column_dimensions['B'].width = 15
ws_schedule.column_dimensions['C'].width = 12
ws_schedule.column_dimensions['D'].width = 30
ws_schedule.column_dimensions['E'].width = 12
ws_schedule.column_dimensions['F'].width = 12
ws_schedule.column_dimensions['G'].width = 12
ws_schedule.column_dimensions['H'].width = 12

# Create Holidays sheet
ws_holidays = wb.create_sheet('Holidays')
ws_holidays['A1'] = '공휴일'
ws_holidays['A1'].font = Font(bold=True, color='FFFFFF')
ws_holidays['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
ws_holidays.column_dimensions['A'].width = 15

# Add 2026 holidays (Korea)
holidays_2026 = [
    '2026-01-01',  # New Year
    '2026-03-01',  # Independence Movement Day
    '2026-05-05',  # Children's Day
    '2026-06-06',  # Memorial Day
    '2026-08-15',  # Liberation Day
    '2026-10-03',  # National Foundation Day
    '2026-10-09',  # Hangeul Day
    '2026-12-25',  # Christmas
]

for i, holiday in enumerate(holidays_2026, 2):
    ws_holidays[f'A{i}'] = holiday

# Generate sample data
workers = ['작업자1', '작업자2', '작업자3']
categories = ['설계', '개발', '테스트', '배포']
sub_categories = ['화면', 'DB', '인터페이스', 'API']
task_names = ['로그인기능', '메인화면', '데이터조회', '보고서출력', '승인처리', 
              '통계분석', '파일업로드', '데이터검증', '알림기능', '설정관리']

current_row = 2
start_date = datetime(2026, 3, 3)
end_date = datetime(2026, 5, 31)

# Generate 10 tasks per worker
for worker in workers:
    worker_tasks = []
    current_task_date = start_date
    
    for task_num in range(10):
        # Random task duration (2-8 working days)
        working_days = random.randint(2, 8)
        
        task_start = current_task_date
        # Calculate end date considering weekends
        days_added = 0
        temp_date = task_start
        while days_added < working_days:
            if temp_date.weekday() < 5:  # Monday=0, Friday=4
                days_added += 1
            if days_added < working_days:
                temp_date += timedelta(days=1)
        task_end = temp_date
        
        # Move to next task start (day after current end)
        current_task_date = task_end + timedelta(days=1)
        # Skip weekends
        while current_task_date.weekday() >= 5:
            current_task_date += timedelta(days=1)
        
        worker_tasks.append({
            'category': random.choice(categories),
            'sub_category': random.choice(sub_categories),
            'program_id': f'PRG{task_num+1:03d}',
            'task_name': random.choice(task_names),
            'worker': worker,
            'start_date': task_start.strftime('%Y-%m-%d'),
            'end_date': task_end.strftime('%Y-%m-%d')
        })
    
    # Shuffle tasks to randomize order
    random.shuffle(worker_tasks)
    
    # Write to sheet
    for task in worker_tasks:
        ws_schedule[f'A{current_row}'] = task['category']
        ws_schedule[f'B{current_row}'] = task['sub_category']
        ws_schedule[f'C{current_row}'] = task['program_id']
        ws_schedule[f'D{current_row}'] = task['task_name']
        ws_schedule[f'E{current_row}'] = task['worker']
        ws_schedule[f'F{current_row}'] = task['start_date']
        ws_schedule[f'G{current_row}'] = task['end_date']
        
        # NETWORKDAYS formula
        formula = f'=NETWORKDAYS(F{current_row},G{current_row},Holidays!$A:$A)'
        ws_schedule[f'H{current_row}'] = formula
        
        current_row += 1

print(f'Generated {current_row-2} tasks')

# Create 누락날짜 sheet
ws_missing = wb.create_sheet('누락날짜시트')

# Headers for missing dates sheet
ws_missing['A1'] = '작업자'
ws_missing['A1'].font = Font(bold=True, color='FFFFFF')
ws_missing['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
ws_missing['A1'].alignment = Alignment(horizontal='center', vertical='center')

ws_missing['B1'] = '누락된 날짜'
ws_missing['B1'].font = Font(bold=True, color='FFFFFF')
ws_missing['B1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
ws_missing['B1'].alignment = Alignment(horizontal='center', vertical='center')

ws_missing.column_dimensions['A'].width = 12
ws_missing.column_dimensions['B'].width = 15

# Add instructions
ws_missing['A3'] = '※ 작업자 입력:'
ws_missing['B3'] = '작업자1'
ws_missing['B3'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

ws_missing['A5'] = '※ 사용 수식 예시:'
ws_missing['A6'] = '=MINIFS(일정관리!$F:$F,일정관리!$E:$E,$B$3)'
ws_missing['A7'] = '(작업자의 최소 시작일)'
ws_missing['A8'] = '=MAXIFS(일정관리!$G:$G,일정관리!$E:$E,$B$3)'
ws_missing['A9'] = '(작업자의 최대 종료일)'

# Save workbook
wb.save(r'D:\workspace\schedule\프로젝트일정관리_템플릿.xlsx')
print('Excel file created successfully at D:\\workspace\\schedule\\프로젝트일정관리_템플릿.xlsx')
